from io import BytesIO
import pandas as pd

from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import (
    Font,
    Alignment,
    PatternFill,
    Border,
    Side
)
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


def create_excel_export(
    product_type,
    tenor_months,
    observation_frequency,
    first_call_month,
    autocall_trigger,
    step_down_size,
    income_trigger,
    memory_coupon,
    coupon_pa,
    capital_barrier,
    notional,
    date_column,
    price_columns,
    results,
    summary_stats,
    autocall_summary=None,
    underlying_performance_data=None,
    participation_rate=None,
    participation_strike=None,
    protection_type=None,
    protection_level=None,
    upside_cap=None
):
    output = BytesIO()

    # =========================
    # Inputs
    # =========================

    if product_type == "Participation":

        inputs_df = pd.DataFrame({
            "Input": [
                "Product Type",
                "Tenor months",
                "Participation Rate (%)",
                "Participation Strike (%)",
                "Protection Type",
                "Protection Level (%)",
                "Upside Cap",
                "Date Column",
                "Underlying Columns"
            ],
            "Value": [
                product_type,
                tenor_months,
                participation_rate,
                participation_strike,
                protection_type,
                protection_level,
                (
                    upside_cap
                    if upside_cap is not None
                    else "None"
                ),
                date_column,
                ", ".join(price_columns)
            ]
        })

    else:

        inputs_df = pd.DataFrame({
            "Input": [
                "Product Type",
                "Tenor months",
                "Observation Frequency",
                "First Call Month",
                "Autocall Trigger (%)",
                "Step-Down Size (%)",
                "Income Trigger (%)",
                "Memory Coupon",
                "Coupon p.a. (%)",
                "Capital Barrier (%)",
                "Date Column",
                "Underlying Columns"
            ],
            "Value": [
                product_type,
                tenor_months,
                observation_frequency,
                first_call_month,
                autocall_trigger,
                step_down_size,
                income_trigger,
                memory_coupon,
                coupon_pa,
                capital_barrier,
                date_column,
                ", ".join(price_columns)
            ]
        })

    with pd.ExcelWriter(
        output,
        engine="openpyxl"
    ) as writer:

        # =========================
        # Core sheets
        # =========================

        inputs_df.to_excel(
            writer,
            sheet_name="Inputs",
            index=False
        )

        results.to_excel(
            writer,
            sheet_name="Backtest Results",
            index=False
        )

        summary_stats.to_excel(
            writer,
            sheet_name="Summary",
            index=False
        )

        # =========================
        # Autocall Distribution
        # =========================

        if autocall_summary is not None:

            autocall_export = (
                autocall_summary.copy()
            )

            autocall_export["%"] = (
                autocall_export["%"]
                .astype(str)
                .str.replace(
                    "%",
                    "",
                    regex=False
                )
                .astype(float)
                / 100
            )

            autocall_export[
                "Chart Month"
            ] = (
                autocall_export[
                    "Autocall Test"
                ]
                .str.replace(
                    " Months",
                    "",
                    regex=False
                )
            )

            autocall_export.loc[
                autocall_export[
                    "Autocall Test"
                ].isin(
                    [
                        "Autocall Missed",
                        "Total"
                    ]
                ),
                "Chart Month"
            ] = None

            autocall_export.to_excel(
                writer,
                sheet_name="Autocall Distribution",
                index=False
            )

        # =========================
        # Phoenix Coupon Analysis
        # =========================

        if product_type in [
            "Phoenix Autocall",
            "Step-Down Phoenix Autocall"
        ]:

            phoenix_columns = [
                col
                for col in results.columns
                if col in [
                    "Coupon Paid This Observation (%)",
                    "Missed Coupon Bank (%)",
                    "Coupon Paid on Final Observation (%)",
                    "Total Coupons Paid (%)",
                    "Coupons Paid",
                    "Coupon Opportunities Until Exit",
                    "Coupon Capture Rate (%)"
                ]
            ]

            if phoenix_columns:

                results[
                    phoenix_columns
                ].to_excel(
                    writer,
                    sheet_name="Phoenix Coupon Analysis",
                    index=False
                )

        # =========================
        # Underlying Performance Data
        # =========================

        if (
            underlying_performance_data
            is not None
        ):

            performance_export = (
                underlying_performance_data.copy()
            )

            for col in price_columns:

                if (
                    col
                    in performance_export.columns
                ):

                    performance_export[col] = (
                        performance_export[col]
                        - 100
                    )

            performance_export.to_excel(
                writer,
                sheet_name="Underlying Performance Data",
                index=False
            )

        # =========================
        # Workbook formatting
        # =========================

        workbook = writer.book

        if (
            "Autocall Distribution"
            in workbook.sheetnames
        ):

            autocall_sheet = workbook[
                "Autocall Distribution"
            ]

            for cell in (
                autocall_sheet["C"][1:]
            ):
                cell.number_format = "0.00%"

            autocall_sheet.column_dimensions[
                "D"
            ].hidden = True

        for worksheet in workbook.worksheets:

            worksheet.sheet_view.showGridLines = False

            if worksheet.max_row >= 1:

                for cell in worksheet[1]:

                    cell.font = Font(
                        bold=True
                    )

        # =========================
        # Backtest Results
        # Column Explanations
        # =========================

        if "Backtest Results" in workbook.sheetnames:

            results_sheet = workbook[
                "Backtest Results"
            ]

            column_explanations = {

                # General
                "Trade Date": (
                    "Historical start date of the simulated investment."
                ),

                "Maturity Date": (
                    "Date used as the maturity observation for the "
                    "simulated investment."
                ),

                "Observation Month": (
                    "Number of months from the Trade Date to the "
                    "observation at which the product exited or matured."
                ),

                "Observation Year": (
                    "Observation Month divided by 12."
                ),

                "Event": (
                    "Outcome of the simulated investment, such as "
                    "Autocalled, capital protected, barrier breached "
                    "or participation payoff."
                ),

                # Underlying calculations
                "Worst Underlying": (
                    "The selected underlying with the lowest performance "
                    "relative to its initial level."
                ),

                "Worst Initial Level": (
                    "Initial level of the worst-performing underlying "
                    "on the Trade Date."
                ),

                "Worst Final Level": (
                    "Final observed level of the worst-performing "
                    "underlying."
                ),

                "Worst Performance (%)": (
                    "Final level of the worst-performing underlying "
                    "divided by its initial level, expressed as a "
                    "percentage."
                ),

                "Underlying Level (%)": (
                    "Final level of the worst-performing underlying "
                    "expressed as a percentage of its initial level."
                ),

                "Underlying Performance (%)": (
                    "Percentage change in the worst-performing underlying "
                    "from its initial level."
                ),

                # Autocall
                "Autocall Trigger (%)": (
                    "Required underlying level for the product to "
                    "autocall at the relevant observation."
                ),

                # Phoenix
                "Coupon Paid This Observation (%)": (
                    "Income coupon paid at the relevant observation."
                ),

                "Missed Coupon Bank (%)": (
                    "Unpaid coupons accumulated under the memory coupon "
                    "feature and available to be recovered later."
                ),

                "Coupon Paid on Final Observation (%)": (
                    "Income coupon paid on the final observation date."
                ),

                "Total Coupons Paid (%)": (
                    "Total income coupons paid over the life of the "
                    "simulated product."
                ),

                "Coupons Paid": (
                    "Number of individual scheduled income payments "
                    "received during the simulation. For a memory coupon, "
                    "recovered missed coupons are included."
                ),

                "Coupon Opportunities Until Exit": (
                    "Number of scheduled income observation opportunities "
                    "from the Trade Date until autocall or maturity."
                ),

                "Coupon Capture Rate (%)": (
                    "Coupons Paid divided by Coupon Opportunities Until "
                    "Exit."
                ),

                # Participation
                "Protection Type": (
                    "Selected downside protection structure for the "
                    "Participation product."
                ),

                "Protection Level (%)": (
                    "Capital protection or minimum payoff level used by "
                    "the Participation structure."
                ),

                "Participation Rate (%)": (
                    "Percentage participation applied to the positive "
                    "return above the Participation Strike."
                ),

                "Participation Strike (%)": (
                    "Underlying level from which positive participation "
                    "begins."
                ),

                "Call Return (%)": (
                    "Maximum of Underlying Level minus Participation "
                    "Strike and zero."
                ),

                "Participation Return (%)": (
                    "Call Return multiplied by the Participation Rate."
                ),

                "Upside Cap": (
                    "Maximum payoff level where an upside cap has been "
                    "selected. None means the payoff is uncapped."
                ),

                # Return calculations
                "Return (%)": (
                    "Total investment return calculated as Payoff minus "
                    "the initial reference amount of 100."
                ),

                "Payoff": (
                    "Final redemption value based on an initial reference "
                    "amount of 100, including coupons, participation and "
                    "any capital loss where applicable."
                ),

                "Flat Coupon Return p.a. (%)": (
                    "Total return divided by the number of years the "
                    "investment was outstanding. This is a simple, "
                    "non-compounded annual return."
                ),

                "Annualised Return (%)": (
                    "Compound annual growth rate based on the final payoff "
                    "and the actual time the simulated investment was "
                    "outstanding."
                )
            }

            explanation_start_col = (
                results_sheet.max_column + 3
            )

            title_cell = results_sheet.cell(
                row=1,
                column=explanation_start_col
            )

            title_cell.value = (
                "Column Explanations"
            )

            title_cell.font = Font(
                bold=True,
                size=12
            )

            title_cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9EAF7"
            )

            results_sheet.merge_cells(
                start_row=1,
                start_column=explanation_start_col,
                end_row=1,
                end_column=explanation_start_col + 1
            )

            column_header = results_sheet.cell(
                row=2,
                column=explanation_start_col
            )

            explanation_header = results_sheet.cell(
                row=2,
                column=explanation_start_col + 1
            )

            column_header.value = "Column"

            explanation_header.value = (
                "Calculation / Explanation"
            )

            for cell in [
                column_header,
                explanation_header
            ]:

                cell.font = Font(
                    bold=True
                )

                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor="E7E6E6"
                )

            thin_border = Border(
                left=Side(
                    style="thin",
                    color="BFBFBF"
                ),
                right=Side(
                    style="thin",
                    color="BFBFBF"
                ),
                top=Side(
                    style="thin",
                    color="BFBFBF"
                ),
                bottom=Side(
                    style="thin",
                    color="BFBFBF"
                )
            )

            explanation_row = 3

            for column_name in results.columns:

                if (
                    column_name
                    not in column_explanations
                ):
                    continue

                name_cell = results_sheet.cell(
                    row=explanation_row,
                    column=explanation_start_col
                )

                text_cell = results_sheet.cell(
                    row=explanation_row,
                    column=explanation_start_col + 1
                )

                name_cell.value = (
                    column_name
                )

                text_cell.value = (
                    column_explanations[
                        column_name
                    ]
                )

                name_cell.font = Font(
                    bold=True
                )

                name_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                text_cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True
                )

                name_cell.border = thin_border
                text_cell.border = thin_border

                explanation_row += 1

            # Use get_column_letter here because
            # the title cells above are merged
            explanation_name_col = (
                get_column_letter(
                    explanation_start_col
                )
            )

            explanation_text_col = (
                get_column_letter(
                    explanation_start_col + 1
                )
            )

            results_sheet.column_dimensions[
                explanation_name_col
            ].width = 34

            results_sheet.column_dimensions[
                explanation_text_col
            ].width = 70

        # =========================
        # Charts sheet
        # =========================

        charts_sheet = workbook.create_sheet(
            "Charts"
        )

        charts_sheet.sheet_view.showGridLines = False

        charts_sheet["A1"] = " "

        # =========================
        # Native Excel Autocall Chart
        # =========================

        if (
            autocall_summary is not None
            and "Autocall Distribution"
            in workbook.sheetnames
        ):

            autocall_sheet = workbook[
                "Autocall Distribution"
            ]

            chart_last_row = (
                autocall_sheet.max_row
                - 2
            )

            if chart_last_row >= 2:

                autocall_chart = (
                    BarChart()
                )

                autocall_chart.type = (
                    "col"
                )

                autocall_chart.title = (
                    "Autocall Back-Test"
                )

                autocall_chart.height = 12
                autocall_chart.width = 24

                autocall_chart.legend = None

                autocall_chart.x_axis.title = (
                    "Observation Point in "
                    "Months from Strike"
                )

                autocall_chart.y_axis.title = ""

                autocall_chart.y_axis.numFmt = (
                    "0.00%"
                )

                autocall_chart\
                    .y_axis\
                    .scaling\
                    .min = 0

                data = Reference(
                    autocall_sheet,
                    min_col=3,
                    min_row=1,
                    max_row=chart_last_row
                )

                categories = Reference(
                    autocall_sheet,
                    min_col=4,
                    min_row=2,
                    max_row=chart_last_row
                )

                autocall_chart.add_data(
                    data,
                    titles_from_data=True
                )

                autocall_chart.set_categories(
                    categories
                )

                series = (
                    autocall_chart.series[0]
                )

                series\
                    .graphicalProperties\
                    .solidFill = "2F75B5"

                series\
                    .graphicalProperties\
                    .line\
                    .solidFill = "2F75B5"

                autocall_chart.gapWidth = 180

                autocall_chart.dLbls = (
                    DataLabelList()
                )

                autocall_chart\
                    .dLbls\
                    .showVal = True

                autocall_chart\
                    .dLbls\
                    .numFmt = "0.00%"

                autocall_chart\
                    .dLbls\
                    .position = "outEnd"

                charts_sheet.add_chart(
                    autocall_chart,
                    "A3"
                )

        # =========================
        # Native Excel Underlying Chart
        # =========================

        if (
            underlying_performance_data
            is not None
            and "Underlying Performance Data"
            in workbook.sheetnames
        ):

            performance_sheet = workbook[
                "Underlying Performance Data"
            ]

            if (
                performance_sheet.max_row >= 2
                and
                performance_sheet.max_column >= 2
            ):

                underlying_chart = (
                    LineChart()
                )

                underlying_chart.title = ""

                underlying_chart.height = 14
                underlying_chart.width = 24

                underlying_chart\
                    .legend\
                    .position = "t"

                underlying_chart\
                    .y_axis\
                    .title = ""

                underlying_chart\
                    .x_axis\
                    .title = ""

                underlying_chart\
                    .y_axis\
                    .numFmt = '0"%"'

                underlying_chart\
                    .y_axis\
                    .scaling\
                    .min = None

                underlying_chart\
                    .y_axis\
                    .scaling\
                    .max = None

                data = Reference(
                    performance_sheet,
                    min_col=2,
                    min_row=1,
                    max_col=(
                        performance_sheet
                        .max_column
                    ),
                    max_row=(
                        performance_sheet
                        .max_row
                    )
                )

                dates = Reference(
                    performance_sheet,
                    min_col=1,
                    min_row=2,
                    max_row=(
                        performance_sheet
                        .max_row
                    )
                )

                underlying_chart.add_data(
                    data,
                    titles_from_data=True,
                    from_rows=False
                )

                underlying_chart.set_categories(
                    dates
                )

                line_colours = [
                    "548235",
                    "FF5B57",
                    "F4B183",
                    "1F3A93",
                    "70AD47"
                ]

                for i, series in enumerate(
                    underlying_chart.series
                ):

                    colour = line_colours[
                        i % len(
                            line_colours
                        )
                    ]

                    series\
                        .graphicalProperties\
                        .line\
                        .solidFill = colour

                    series\
                        .graphicalProperties\
                        .line\
                        .width = 18000

                    series.marker.symbol = (
                        "none"
                    )

                underlying_chart\
                    .x_axis\
                    .number_format = (
                        "dd/mm/yy"
                    )

                charts_sheet.add_chart(
                    underlying_chart,
                    "A28"
                )

        # =========================
        # Basic column widths
        # =========================

        for sheet_name in [
            "Inputs",
            "Summary",
            "Autocall Distribution",
            "Phoenix Coupon Analysis"
        ]:

            if (
                sheet_name
                in workbook.sheetnames
            ):

                ws = workbook[
                    sheet_name
                ]

                for column_cells in (
                    ws.columns
                ):

                    max_length = 0

                    column_letter = (
                        column_cells[0]
                        .column_letter
                    )

                    for cell in column_cells:

                        try:

                            if (
                                cell.value
                                is not None
                            ):

                                max_length = max(
                                    max_length,
                                    len(
                                        str(
                                            cell.value
                                        )
                                    )
                                )

                        except Exception:
                            pass

                    ws.column_dimensions[
                        column_letter
                    ].width = min(
                        max_length + 2,
                        35
                    )

        # =========================
        # Underlying data date format
        # =========================

        if (
            "Underlying Performance Data"
            in workbook.sheetnames
        ):

            performance_sheet = workbook[
                "Underlying Performance Data"
            ]

            for cell in (
                performance_sheet["A"][1:]
            ):

                cell.number_format = (
                    "dd/mm/yyyy"
                )

    output.seek(0)

    return output
